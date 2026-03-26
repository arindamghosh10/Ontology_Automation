"""
Excel file handler for the Ontology Automation Agent.
Reads merchant records and writes results back using openpyxl.
"""

import logging
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Expected columns in Sheet1
SHEET1_COLUMNS = [
    "store_id", "store_name", "store_domain",
    "corrected_website_text", "corrected_merchant_zoominfo_url",
    "corrected_zoominfo_page_text", "corrected_merchant_dnb_url",
    "corrected_dnb_page_text", "corrected_merchant_wikipedia_url",
    "corrected_wikipedia_page_text", "acquisitions",
    "other_phone_numbers", "verification_notes", "Confidence_Score"
]

LOCATION_COLUMNS = [
    "store_name", "store_id", "Address", "Phone Number"
]

REVIEW_COLUMNS = [
    "store_id", "store_name", "field_name", "candidate_value",
    "reason_flagged", "confidence_score", "recommended_action"
]


def _get_or_create_sheet(wb, sheet_name: str, columns: list[str]):
    """Get existing sheet or create new one with headers."""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    return ws


def _find_column_index(ws, column_name: str) -> Optional[int]:
    """Find the column index (1-based) for a given column name in row 1."""
    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col_idx).value
        if cell_val and str(cell_val).strip().lower() == column_name.strip().lower():
            return col_idx
    return None


def _build_column_map(ws) -> dict[str, int]:
    """Build a mapping of column_name -> column_index for the header row."""
    col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col_idx).value
        if cell_val:
            col_map[str(cell_val).strip().lower()] = col_idx
    return col_map


def read_merchants(filepath: str, sheet_name: str = "Sheet1") -> list[dict]:
    """
    Read all merchant records from the specified sheet.
    Returns a list of dicts with column names as keys, plus '_row' for the Excel row number.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    wb = load_workbook(str(filepath), read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    col_map = _build_column_map(ws)

    merchants = []
    for row_idx in range(2, ws.max_row + 1):
        # Skip empty rows
        store_id_col = col_map.get("store_id")
        if store_id_col:
            val = ws.cell(row=row_idx, column=store_id_col).value
            if val is None or str(val).strip() == "":
                continue

        merchant = {"_row": row_idx}
        for col_name, col_idx in col_map.items():
            merchant[col_name] = ws.cell(row=row_idx, column=col_idx).value
        merchants.append(merchant)

    wb.close()
    logger.info(f"Read {len(merchants)} merchants from '{sheet_name}'")
    return merchants


def write_merchant_result(filepath: str, row: int, data: dict[str, Any],
                          sheet_name: str = "Sheet1"):
    """
    Write enrichment results for a single merchant to the specified row.
    Only writes columns that exist in 'data'. Creates missing columns if needed.
    """
    filepath = Path(filepath)
    wb = load_workbook(str(filepath))
    ws = wb[sheet_name]
    col_map = _build_column_map(ws)

    for col_name, value in data.items():
        if col_name.startswith("_"):
            continue

        col_name_lower = col_name.strip().lower()
        col_idx = col_map.get(col_name_lower)

        # Create column if it doesn't exist
        if col_idx is None:
            col_idx = ws.max_column + 1
            ws.cell(row=1, column=col_idx, value=col_name)
            col_map[col_name_lower] = col_idx

        # Truncate very long text to avoid Excel cell limits (32767 chars)
        if isinstance(value, str) and len(value) > 32000:
            value = value[:32000] + "\n[TRUNCATED]"

        ws.cell(row=row, column=col_idx, value=value)

    wb.save(str(filepath))
    wb.close()
    logger.info(f"Wrote results for row {row} to '{sheet_name}'")


def write_location_rows(filepath: str, store_name: str, store_id: str,
                        locations: list[dict], sheet_name: str = "Location"):
    """
    Write store location entries to the Location sheet.
    Each location gets its own row: store_name, store_id, Address, Phone Number.
    """
    filepath = Path(filepath)
    wb = load_workbook(str(filepath))
    ws = _get_or_create_sheet(wb, sheet_name, LOCATION_COLUMNS)
    col_map = _build_column_map(ws)

    # Find next empty row
    next_row = ws.max_row + 1

    for loc in locations:
        name_col = col_map.get("store_name", 1)
        id_col = col_map.get("store_id", 2)
        addr_col = col_map.get("address", 3)
        phone_col = col_map.get("phone number", 4)

        ws.cell(row=next_row, column=name_col, value=store_name)
        ws.cell(row=next_row, column=id_col, value=store_id)
        ws.cell(row=next_row, column=addr_col, value=loc.get("address", ""))
        ws.cell(row=next_row, column=phone_col, value=loc.get("phone", ""))
        next_row += 1

    wb.save(str(filepath))
    wb.close()
    logger.info(f"Wrote {len(locations)} locations for '{store_name}' to '{sheet_name}'")


def ensure_review_sheet(filepath: str) -> None:
    """Create the REVIEW sheet with headers if it doesn't exist."""
    filepath = Path(filepath)
    wb = load_workbook(str(filepath))
    _get_or_create_sheet(wb, "REVIEW", REVIEW_COLUMNS)
    wb.save(str(filepath))
    wb.close()


def write_review_entry(filepath: str, entry: dict):
    """
    Append a single review entry to the REVIEW sheet.
    entry should have keys: store_id, store_name, field_name,
    candidate_value, reason_flagged, confidence_score, recommended_action
    """
    filepath = Path(filepath)
    wb = load_workbook(str(filepath))
    ws = _get_or_create_sheet(wb, "REVIEW", REVIEW_COLUMNS)
    col_map = _build_column_map(ws)

    next_row = ws.max_row + 1

    for col_name in REVIEW_COLUMNS:
        col_name_lower = col_name.strip().lower()
        col_idx = col_map.get(col_name_lower)
        if col_idx:
            value = entry.get(col_name_lower) or entry.get(col_name, "")
            if isinstance(value, str) and len(value) > 32000:
                value = value[:32000] + "\n[TRUNCATED]"
            ws.cell(row=next_row, column=col_idx, value=value)

    wb.save(str(filepath))
    wb.close()
    logger.info(f"Wrote review entry for store_id={entry.get('store_id')}, field={entry.get('field_name')}")


def clear_merchant_row(filepath: str, row: int, columns_to_clear: list[str],
                       sheet_name: str = "Sheet1"):
    """Clear specific cells for a merchant row (used when score < 60)."""
    filepath = Path(filepath)
    wb = load_workbook(str(filepath))
    ws = wb[sheet_name]
    col_map = _build_column_map(ws)

    for col_name in columns_to_clear:
        col_idx = col_map.get(col_name.strip().lower())
        if col_idx:
            ws.cell(row=row, column=col_idx, value=None)

    wb.save(str(filepath))
    wb.close()
